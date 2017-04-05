# coding=utf-8 #
from openpyxl import load_workbook
import codecs
import sys
import os

column_separator = '|'
row_separator = '\n'
output_encoding = 'utf-8'
input_ext = '.xlsx'
output_folder = 'output'


def read_workbook(path, name):
    file_name = '%s/%s%s' % (path, name, input_ext)
    print 'process xls: %s' % (file_name)
    workbook = load_workbook(file_name, True, False, True, True)
    sheet_names = workbook.get_sheet_names()
    for sheet_name in sheet_names:
        if not sheet_name.startswith('_'):
            sheet = workbook.get_sheet_by_name(sheet_name)
            output_name = get_output_name(name, sheet_name)
            print 'output csv: %s' % (output_name)
            read_sheet_to(sheet, output_name)


def get_output_name(workbook_name, worksheet_name):
	return '%s/%s.csv' % (output_folder, worksheet_name)
    #return '%s/%s_%s.csv' % (output_folder, workbook_name, worksheet_name)


def read_sheet_to(sheet, filename):
    output_file = codecs.open(filename, 'w', output_encoding)
    current_cell_value = ""
    Current_count_row = 0
    Current_count_line = 0
    arrSkip = []
    for row in sheet.iter_rows():
        for cell in row:
            value = cell.internal_value
            if not isinstance(value, unicode):
                try:
                    current_cell_value = str(int(value))
                except:
                    current_cell_value = str(0)
            else:
                current_cell_value = value
            if Current_count_row == 1:
                if current_cell_value.startswith('_'):
                    arrSkip.append(0)
                else:
                    arrSkip.append(1)
        Current_count_row = Current_count_row + 1
    for row in sheet.iter_rows():
        Current_count_line = 0
        myint = 1
        for cell in row:
            if arrSkip[Current_count_line] == 0:
                continue
            if current_cell_value != "" and myint == 2:
                output_file.write(column_separator)
            myint = 2
            value = cell.internal_value
            if not isinstance(value, unicode):
                try:
                    current_cell_value = str(int(value))
                except:
                    current_cell_value = str(0)
            else:
                current_cell_value = value
            output_file.write(current_cell_value)
            Current_count_line = Current_count_line + 1
        output_file.write(row_separator)
        current_cell_value = ""

    output_file.close()


def process_target(target):
    if os.path.isfile(target):
        process_file(target)

    if os.path.isdir(target):
        process_dir(target)


def process_dir(target):
    for root, dirs, files in os.walk(target):
        for each_file in files:
            if not each_file.startswith('~'):
                file_name = '%s/%s' % (root, each_file)
                process_file(file_name)


def process_file(target):
    (path_name, file_name) = os.path.split(target)
    (base_name, ext) = os.path.splitext(file_name)
    if ext == input_ext:
        read_workbook(path_name, base_name)


def main(targets):
    if not len(targets) == 0:
        for target in targets:
            process_target(target)
    else:
        print 'need arguments for target'


if __name__ == '__main__':
    main(sys.argv[1:])
